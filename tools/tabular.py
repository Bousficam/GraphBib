#!/usr/bin/env python3
"""Shared XLSX/CSV table I/O for GraphBib tools.

XLSX is the default format for every persisted table on the
Extractor side (dedup.xlsx, slug-renames.xlsx, tiab-decisions.xlsx,
fulltext-decisions.xlsx, prisma-counts.xlsx). CSV stays available
as an on-demand export - handy for pandas/R/grep workflows and as a
migration fallback for legacy projects.

Public API:

    read_records(path)
        Return (headers, list[dict]). Auto-detects format by suffix.
        If `path` ends with .xlsx but the file is missing, falls
        back to the sibling .csv - supports projects upgraded from
        the pre-xlsx layout.

    write_records(records, headers, path, *, styled=True)
        Writes xlsx (with bold header, frozen first row, autofilter,
        auto-fit column widths) when path ends with .xlsx.
        Writes CSV otherwise.

    append_record(record, headers, path)
        Add one row at the bottom. For xlsx, loads the workbook,
        appends, saves once. For csv, appends in text mode.
        If the file doesn't exist, creates it with the header row.

    to_csv(xlsx_path, csv_path=None)
        Export the xlsx to a sibling .csv (or to the explicit path).
        Returns the csv path.

CLI:

    python tools/tabular.py to-csv path/to/file.xlsx [output.csv]
    python tools/tabular.py show path/to/file.xlsx [--limit 20]
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


HEADER_FONT = ("Font", "PatternFill")  # placeholder so type checkers see them


def _require_openpyxl():
    if not HAS_XLSX:
        sys.exit("openpyxl not installed. Run: pip install openpyxl")


def detect_format(path):
    """Return 'xlsx' or 'csv' from the suffix. Exits on anything else."""
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".csv":
        return "csv"
    sys.exit(f"Unsupported format: {suffix}. Use .xlsx or .csv.")


def _read_csv_with_encoding_fallback(path):
    """Open CSV trying common encodings - Excel-on-Windows defaults to cp1252."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                reader = csv.DictReader(f)
                headers = list(reader.fieldnames or [])
                data = list(reader)
            if encoding != "utf-8-sig":
                print(
                    f"  · {Path(path).name}: decoded as {encoding} "
                    f"(re-save as UTF-8 for cleaner future reads)",
                    file=sys.stderr,
                )
            return headers, data
        except UnicodeDecodeError:
            continue
    sys.exit(f"Could not decode {path} with utf-8 / cp1252 / latin-1.")


def _read_xlsx(path):
    """Return (headers, list[dict])."""
    _require_openpyxl()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        # Pad short rows with "" so all cells are addressable by header.
        padded = list(r) + [None] * (len(headers) - len(r))
        data.append({h: ("" if v is None else v) for h, v in zip(headers, padded)})
    return headers, data


def read_records(path):
    """Read xlsx or csv (auto-detect). Returns (headers, list[dict]).

    Backward-compat: if `path` ends with .xlsx but the file is
    missing AND a sibling .csv exists, transparently reads the .csv
    (so projects upgraded from the pre-xlsx layout keep working
    until the user writes back as xlsx).
    """
    p = Path(path)
    fmt = detect_format(p)
    if fmt == "xlsx" and not p.exists():
        legacy_csv = p.with_suffix(".csv")
        if legacy_csv.exists():
            print(
                f"  · {p.name} missing, reading legacy {legacy_csv.name} "
                f"(will be re-saved as .xlsx on next write)",
                file=sys.stderr,
            )
            return _read_csv_with_encoding_fallback(legacy_csv)
        # Fall through to the empty case
    if fmt == "xlsx":
        if not p.exists():
            return [], []
        return _read_xlsx(p)
    if not p.exists():
        return [], []
    return _read_csv_with_encoding_fallback(p)


def _write_xlsx(records, headers, path, *, styled=True):
    """Write an xlsx with a styled header, frozen row, autofilter."""
    _require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = Path(path).stem[:31]  # Excel sheet-name cap

    ws.append(headers)
    for row in records:
        ws.append([row.get(h, "") for h in headers])

    if styled and headers:
        # Bold header + light grey fill
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="EFEFEF")
        for cell in ws[1]:
            cell.font = bold
            cell.fill = fill
        # Freeze top row + autofilter on the header range
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Reasonable auto-fit column widths (capped at 60)
        for idx, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for row in records:
                v = row.get(h, "")
                if v is None:
                    continue
                s = str(v)
                if "\n" in s:
                    s = max(s.split("\n"), key=len)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 8), 60)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_csv(records, headers, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in records:
            w.writerow({h: r.get(h, "") for h in headers})


def write_records(records, headers, path, *, styled=True):
    """Write records to xlsx or csv (auto-detect by suffix)."""
    fmt = detect_format(path)
    if fmt == "xlsx":
        _write_xlsx(records, headers, path, styled=styled)
    else:
        _write_csv(records, headers, path)


def append_record(record, headers, path):
    """Append one row to xlsx or csv. Creates the file with header if absent."""
    p = Path(path)
    fmt = detect_format(p)
    if fmt == "xlsx":
        _require_openpyxl()
        if not p.exists():
            _write_xlsx([record], headers, p, styled=True)
            return
        wb = load_workbook(p)
        ws = wb.active
        ws.append([record.get(h, "") for h in headers])
        # Re-extend autofilter to cover the new row
        if ws.auto_filter and ws.dimensions:
            ws.auto_filter.ref = ws.dimensions
        wb.save(p)
        return
    # CSV append
    p.parent.mkdir(parents=True, exist_ok=True)
    write_header = not p.exists()
    with open(p, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        if write_header:
            w.writeheader()
        w.writerow({h: record.get(h, "") for h in headers})


def to_csv(xlsx_path, csv_path=None):
    """Export an xlsx to csv. Returns the csv path."""
    headers, records = _read_xlsx(Path(xlsx_path))
    out = Path(csv_path) if csv_path else Path(xlsx_path).with_suffix(".csv")
    _write_csv(records, headers, out)
    return out


# ----------------------------------------------------------------------
# CLI

def _cli_show(args):
    headers, records = read_records(args.path)
    if not records:
        print(f"(empty: {args.path})")
        return
    print("  ".join(headers))
    print("  ".join("-" * len(h) for h in headers))
    for r in records[: args.limit]:
        print("  ".join(str(r.get(h, ""))[:40] for h in headers))
    if len(records) > args.limit:
        print(f"... ({len(records) - args.limit} more rows)")


def _cli_to_csv(args):
    out = to_csv(args.path, args.output)
    print(f"✓ Wrote {out}")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawTextHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    pshow = sub.add_parser("show", help="Print rows of an xlsx/csv to stdout.")
    pshow.add_argument("path")
    pshow.add_argument("--limit", type=int, default=20)
    pshow.set_defaults(func=_cli_show)

    pcsv = sub.add_parser("to-csv", help="Export xlsx → csv (on-demand).")
    pcsv.add_argument("path", help="Input .xlsx path")
    pcsv.add_argument("output", nargs="?", default=None,
                      help="Output .csv path (defaults to sibling with .csv suffix)")
    pcsv.set_defaults(func=_cli_to_csv)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
