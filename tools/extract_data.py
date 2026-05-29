#!/usr/bin/env python3
"""Data extraction for systematic reviews — fill an Excel/CSV template
from `wiki/sources/`, with a per-column INSTRUCTIONS row and optional
LLM-driven extraction.

Template structure
==================

Row 1:    column headers (e.g. `slug`, `n_intervention`, `risk_of_bias`).
Row 2:    `INSTRUCTIONS` row (slug = "INSTRUCTIONS") — natural-language
          extraction rule per column.
Row 3:    `TYPE` row (slug = "TYPE") — quantitative | ordinal | nominal | text.
Row 4:    `SCALE` row (slug = "SCALE") — for ordinal/nominal: allowed values
          or coded mapping (e.g. "0=low, 1=some concerns, 2=high");
          for quantitative: optional unit hint "(years)", "(0-100)";
          empty for free text.
Row 5+:   one row per source (slug = wiki source basename).

The INSTRUCTIONS / TYPE / SCALE rows are consumed by the tool, validated,
and re-emitted at the top of the output unchanged.

Example:

    | slug         | year      | risk_of_bias                       | design                                          |
    | INSTRUCTIONS | Pub year  | Cochrane RoB 2 overall judgment    | Study design                                    |
    | TYPE         | quantitative | ordinal                         | nominal                                         |
    | SCALE        | (YYYY)    | 0=low, 1=some concerns, 2=high     | RCT, cohort, cross-sectional, case-series       |
    | cervera-2020 |           |                                    |                                                 |
    | khedr-2005   |           |                                    |                                                 |

Filling logic (per cell, in order)
==================================

1. **Frontmatter** (deterministic) — known column names map directly to
   YAML fields (title, authors, year, doi, study_design, sample_size,
   population, intervention_family, etc.).
2. **Body regex** (heuristic) — built-in patterns for clinical fields
   (n per arm, age mean, baseline FM, ΔFM, p-value, Cohen's d, CI,
   trial registration ID, …).
3. **LLM extraction** (with `--llm`) — for cells still empty AND with
   a non-empty INSTRUCTIONS, call Claude (via litellm) with the
   instruction + the source body. Cached in `tools/.cache/extract_llm.json`
   so re-runs are nearly free.

Cells already populated by the user are NEVER overwritten.

Modes
=====

    # Pre-fill a NEW template from a source's cites (default SR columns + INSTRUCTIONS row)
    python tools/extract_data.py --from-source cervera-2020 \\
        --output cervera-extraction.xlsx

    # Fill an existing template (frontmatter + regex only)
    python tools/extract_data.py data_extraction.xlsx

    # Same + LLM for cells with instructions still empty
    python tools/extract_data.py data_extraction.xlsx --llm

    # Override the model
    LLM_MODEL=claude-sonnet-4-5 python tools/extract_data.py data.xlsx --llm

Status reported per row: complete / partial / empty / not_found.
Per-cell method tracked in stderr: frontmatter / regex / llm / manual.
"""
import argparse
import csv
import json
import os
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from _lib import REPO_ROOT, load_sources  # noqa: E402

try:
    from openpyxl import Workbook, load_workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

CACHE_DIR = REPO_ROOT / "tools" / ".cache"
LLM_CACHE_FILE = CACHE_DIR / "extract_llm.json"

INSTRUCTIONS_MARKER = "INSTRUCTIONS"
TYPE_MARKER = "TYPE"
SCALE_MARKER = "SCALE"
SPEC_MARKERS = {INSTRUCTIONS_MARKER, TYPE_MARKER, SCALE_MARKER}

VALID_TYPES = {"quantitative", "ordinal", "nominal", "text"}

# Per-cell LLM extraction is a "grunt-work" task: extract one number /
# one label / one verbatim quote from a known section. Haiku handles
# this well at ~10% the price of Sonnet. Override with LLM_MODEL or
# LLM_MODEL_FAST env vars.
DEFAULT_LLM_MODEL = "claude-haiku-4-5"
LLM_MAX_BODY_CHARS = 60_000   # ~15k tokens
LLM_MAX_TOKENS = 200          # extracted value should be short
LLM_SLEEP_SEC = 0.05          # politeness


# ============================================================================
# Column normalization
# ============================================================================

def normalize_col(col):
    return re.sub(r"[\s\-]+", "_", str(col).lower().strip())


# ============================================================================
# Frontmatter map (deterministic)
# ============================================================================

FM_MAP = {
    "title": "title",
    "authors": lambda s: "; ".join(s["fm"].get("authors") or []),
    "first_author": lambda s: (s["fm"].get("authors") or [""])[0],
    "year": "year",
    "journal": "journal",
    "doi": "doi",
    "design": "study_design",
    "study_design": "study_design",
    "type": "type",
    "sample_size": "sample_size",
    "n": "sample_size",
    "n_total": "sample_size",
    "population": "population",
    "intervention": lambda s: s["fm"].get("intervention_family") or "",
    "intervention_family": "intervention_family",
    "intervention_subfamily": "intervention_subfamily",
    "interventions": lambda s: "; ".join(s["fm"].get("interventions") or []),
    "methods": lambda s: "; ".join(s["fm"].get("methods") or []),
    "domain": lambda s: "; ".join(s["fm"].get("domain") or []),
    "language": "language",
    "peer_reviewed": "peer_reviewed",
    "preprint": "preprint",
    "citation_apa": "citation_apa",
    "bibtex_key": "bibtex_key",
    "ingest_date": "date",
    "source_pdf": "source_pdf",
}


# ============================================================================
# Body regex patterns (heuristic)
# ============================================================================

BODY_PATTERNS = {
    "n_intervention": [
        r"(?:intervention|experimental|active|treatment)\s*(?:group)?\s*[:,]?\s*[Nn]\s*=\s*(\d+)",
        r"(\d+)\s*(?:patients|participants|subjects).{0,40}(?:intervention|experimental|active|treatment)",
    ],
    "n_int": "n_intervention",
    "n_active": "n_intervention",
    "n_experimental": "n_intervention",
    "n_control": [
        r"(?:control|sham|placebo|waitlist)\s*(?:group)?\s*[:,]?\s*[Nn]\s*=\s*(\d+)",
        r"(\d+)\s*(?:patients|participants|subjects).{0,40}(?:control|sham|placebo|waitlist)",
    ],
    "n_ctrl": "n_control",
    "n_sham": "n_control",
    "n_placebo": "n_control",
    "age_mean": [
        r"mean\s*age\s*[=:]?\s*(\d+\.?\d*)",
        r"age\s*(?:was)?\s*[=:]?\s*(\d+\.?\d*)\s*[±+\-]\s*\d+",
        r"age\s*\(M\s*[=:]?\s*(\d+\.?\d*)",
    ],
    "age_sd": [r"age.{0,30}[±+\-]\s*(\d+\.?\d*)"],
    "sex_pct_female": [
        r"(\d+\.?\d*)\s*%\s*(?:female|women)",
        r"(?:female|women).{0,10}(\d+\.?\d*)\s*%",
    ],
    "chronicity": [
        r"(chronic|subacute|acute)\s*stroke\s+patients",
        r"(?:time\s+(?:since|post)\s+stroke|chronicity)\s*[=:]?\s*(\d+\.?\d*\s*(?:months?|weeks?|years?))",
    ],
    "baseline_fm": [
        r"(?:baseline|pre[-\s]?intervention|initial).{0,40}Fugl[-\s]?Meyer.{0,30}?(\d+\.?\d*)",
        r"Fugl[-\s]?Meyer.{0,40}(?:baseline|pre[-\s]?intervention).{0,30}?(\d+\.?\d*)",
    ],
    "baseline_fugl_meyer": "baseline_fm",
    "fm_baseline": "baseline_fm",
    "primary_outcome_delta": [
        r"Δ\s*(?:FM|Fugl[-\s]?Meyer)\s*=\s*([+-]?\d+\.?\d*)",
        r"(?:change|gain|increase).{0,50}(?:FM|Fugl[-\s]?Meyer).{0,30}([+-]?\d+\.?\d*)",
    ],
    "delta_fm": "primary_outcome_delta",
    "primary_delta_fm": "primary_outcome_delta",
    "delta_arat": [
        r"Δ\s*ARAT\s*=\s*([+-]?\d+\.?\d*)",
        r"(?:change|gain).{0,30}ARAT.{0,30}([+-]?\d+\.?\d*)",
    ],
    "p_value": [r"\bp\s*[<≤=]\s*(0?\.\d+|\d+\.\d+)"],
    "p": "p_value",
    "pvalue": "p_value",
    "effect_size": [
        r"(?:Cohen's\s*)?[dgη]\s*[=:]\s*([+-]?\d+\.\d+)",
    ],
    "cohen_d": [
        r"Cohen's\s*d\s*[=:]\s*([+-]?\d+\.\d+)",
        r"\bd\s*[=:]\s*([+-]?\d+\.\d+)",
    ],
    "hedges_g": [
        r"Hedges'?\s*g\s*[=:]\s*([+-]?\d+\.\d+)",
        r"\bg\s*[=:]\s*([+-]?\d+\.\d+)",
    ],
    "confidence_interval": [
        r"95\s*%\s*CI\s*[:\(\[]?\s*([+-]?\d+\.?\d*\s*(?:to|–|-|,)\s*[+-]?\d+\.?\d*)",
    ],
    "ci_95": "confidence_interval",
    "adverse_events": [
        r"(?:adverse events|AE).{0,30}(?:reported|occurred).{0,30}(\d+)",
        r"(no adverse events)",
    ],
    "ae": "adverse_events",
    "trial_registration": [
        r"\b(NCT\d{8})\b",
        r"\b(ISRCTN\d+)\b",
        r"\b(CTRI/\d{4}/\d+/\d+)\b",
        r"\b(EudraCT\s*\d{4}-\d{6}-\d{2})\b",
    ],
    "trial_id": "trial_registration",
    "country": [
        r"(?:was|were)\s+(?:conducted|recruited|performed)\s+in\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)",
    ],
    "n_sessions": [
        r"(\d+)\s+sessions",
        r"(\d+)\s+training\s+sessions",
    ],
    "session_duration_min": [
        r"(\d+)\s*[-–]?\s*(?:min|minute)\s+(?:long\s+)?sessions",
    ],
}


def resolve_pattern(col_norm):
    val = BODY_PATTERNS.get(col_norm)
    seen = set()
    while isinstance(val, str):
        if val in seen:
            return None
        seen.add(val)
        val = BODY_PATTERNS.get(val)
    return val


# ============================================================================
# Section windowing — slice the relevant IMRAD section before LLM call
# ============================================================================

# Map (normalized column → section keyword likely to contain the answer).
# Used by find_relevant_section() to slice the body before sending to LLM.
COLUMN_TO_SECTION = {
    # Methods → Participants
    "n_intervention": "methods",
    "n_int": "methods",
    "n_active": "methods",
    "n_experimental": "methods",
    "n_control": "methods",
    "n_ctrl": "methods",
    "n_sham": "methods",
    "n_placebo": "methods",
    "n_total": "methods",
    "age_mean": "methods",
    "age_sd": "methods",
    "sex_pct_female": "methods",
    "population": "methods",
    "chronicity": "methods",
    "baseline_fm": "methods",
    "intervention": "methods",
    "intervention_subfamily": "methods",
    "intervention_family": "methods",
    "comparator": "methods",
    "n_sessions": "methods",
    "session_duration_min": "methods",
    "trial_registration": "methods",
    "trial_id": "methods",
    "country": "methods",

    # Results
    "primary_outcome_delta": "results",
    "primary_delta_fm": "results",
    "delta_fm": "results",
    "delta_arat": "results",
    "p_value": "results",
    "p": "results",
    "pvalue": "results",
    "effect_size": "results",
    "cohen_d": "results",
    "hedges_g": "results",
    "confidence_interval": "results",
    "ci_95": "results",
    "adverse_events": "results",
    "ae": "results",

    # Discussion
    "limitations": "discussion",
    "comparison_with_prior": "discussion",
    "future_research": "discussion",

    # Background / Introduction
    "background": "introduction",
    "rationale": "introduction",
}

SECTION_HEADINGS = {
    "introduction": [
        r"^##\s+Introduction\b", r"^##\s+Background\b",
        r"^##\s+\d+\.?\s*Introduction\b",
    ],
    "methods": [
        r"^##\s+Methods?\b", r"^##\s+Methodology\b", r"^##\s+Materials?\s+and\s+Methods?\b",
        r"^##\s+Study\s+Design\b", r"^##\s+\d+\.?\s*Methods?\b",
    ],
    "results": [
        r"^##\s+Results?\b", r"^##\s+Findings?\b",
        r"^##\s+\d+\.?\s*Results?\b",
    ],
    "discussion": [
        r"^##\s+Discussion\b", r"^##\s+General\s+Discussion\b",
        r"^##\s+\d+\.?\s*Discussion\b",
    ],
    "conclusion": [
        r"^##\s+Conclusions?\b", r"^##\s+Summary\b",
    ],
    "references": [
        r"^##\s+References\b", r"^##\s+Bibliography\b",
    ],
}


def find_relevant_section(body, section_name, fallback_chars=20_000):
    """Return the slice of `body` corresponding to `section_name`.

    Looks for the section's `## …` heading and slices until the next
    same-level heading (or EOF). If the section isn't found, returns
    the first `fallback_chars` characters of the body so the LLM has
    something to work with.

    Always falls back to the full body if the slice would be too short
    (< 500 chars) — better to send more than to send a stub.
    """
    pats = SECTION_HEADINGS.get(section_name, [])
    if not pats:
        return body[:fallback_chars]
    for pat in pats:
        m = re.search(pat, body, re.MULTILINE | re.IGNORECASE)
        if not m:
            continue
        start = m.start()
        nxt = re.search(r"^##\s+", body[m.end():], re.MULTILINE)
        end = m.end() + nxt.start() if nxt else len(body)
        slice_ = body[start:end]
        if len(slice_) >= 500:
            return slice_
    return body[:fallback_chars]


def window_body(body, col_norm):
    """Return either the relevant IMRAD section for the column, or the full body."""
    section = COLUMN_TO_SECTION.get(col_norm)
    if not section:
        return body, "full"
    sliced = find_relevant_section(body, section)
    if len(sliced) < len(body) * 0.6:
        # Real saving — use the slice
        return sliced, section
    # The slice covers most of the body anyway, send everything
    return body, "full"


def apply_fm_map(col_norm, source):
    spec = FM_MAP.get(col_norm)
    if spec is None:
        return None
    if callable(spec):
        return spec(source)
    val = source["fm"].get(spec)
    if isinstance(val, list):
        return "; ".join(str(v) for v in val)
    return val if val is not None else ""


def extract_from_body(col_norm, body):
    patterns = resolve_pattern(col_norm)
    if not patterns:
        return None
    for pat in patterns:
        m = re.search(pat, body, re.IGNORECASE)
        if m:
            return m.group(1) if m.groups() else m.group(0)
    return ""


# ============================================================================
# LLM extraction (--llm mode)
# ============================================================================

def load_llm_cache():
    if LLM_CACHE_FILE.exists():
        try:
            return json.loads(LLM_CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_llm_cache(cache):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    LLM_CACHE_FILE.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")


def parse_scale(scale_str):
    """Parse a SCALE string into a structured form.

    - "0=low, 1=some concerns, 2=high"   → {"kind": "mapping", "items": [("0","low"), ("1","some concerns"), ("2","high")]}
    - "RCT, cohort, cross-sectional"     → {"kind": "enum", "items": ["RCT", "cohort", "cross-sectional"]}
    - "(years)" or "(0-100)" or empty    → {"kind": "hint", "items": ["(years)"]}  (free text annotation)
    """
    s = (scale_str or "").strip()
    if not s:
        return None
    if "=" in s:
        items = []
        for part in re.split(r"[,;]", s):
            part = part.strip()
            if "=" in part:
                k, v = part.split("=", 1)
                items.append((k.strip(), v.strip()))
            elif part:
                items.append((part, part))
        return {"kind": "mapping", "items": items}
    if s.startswith("(") and s.endswith(")"):
        return {"kind": "hint", "items": [s]}
    if "," in s:
        items = [p.strip() for p in s.split(",") if p.strip()]
        return {"kind": "enum", "items": items}
    return {"kind": "hint", "items": [s]}


def render_scale_for_prompt(scale_dict):
    if not scale_dict:
        return ""
    if scale_dict["kind"] == "mapping":
        pairs = "; ".join(f"{k} = {v}" for k, v in scale_dict["items"])
        return f"Allowed values (return the CODE, not the label): {pairs}"
    if scale_dict["kind"] == "enum":
        vals = ", ".join(scale_dict["items"])
        return f"Allowed values (return one of these verbatim): {vals}"
    if scale_dict["kind"] == "hint":
        return f"Format hint: {scale_dict['items'][0]}"
    return ""


def validate_value(value, type_str, scale_dict):
    """Return (ok, normalized_value, warning_or_None)."""
    v = (value or "").strip()
    if not v:
        return True, "", None
    if v.lower() == "not reported":
        return True, "not reported", None
    if v.startswith("<llm-error"):
        return False, v, "llm-error"

    t = (type_str or "").strip().lower()

    if t == "quantitative":
        if re.search(r"\d", v):
            return True, v, None
        return False, v, "expected a number"

    if t in ("ordinal", "nominal") and scale_dict:
        if scale_dict["kind"] == "mapping":
            codes = [k.lower() for k, _ in scale_dict["items"]]
            labels = [lbl.lower() for _, lbl in scale_dict["items"]]
            v_lc = v.lower()
            # Match a code → keep as code
            for k, _ in scale_dict["items"]:
                if v_lc == k.lower():
                    return True, k, None
            # Match a label → return its code
            for k, lbl in scale_dict["items"]:
                if v_lc == lbl.lower():
                    return True, k, None
            return False, v, f"not in scale codes={codes} or labels={labels}"
        if scale_dict["kind"] == "enum":
            v_lc = v.lower()
            for opt in scale_dict["items"]:
                if v_lc == opt.lower():
                    return True, opt, None
            return False, v, f"not in allowed values {scale_dict['items']}"

    # text or no scale → accept as-is
    return True, v, None


def llm_extract(column_name, instruction, body, type_str=None, scale_dict=None):
    """Single-cell LLM extraction with prompt caching on the body.

    The static system prompt and the paper body are sent as separate
    content blocks with cache_control: ephemeral. When multiple cells
    of the SAME paper are processed in sequence, calls 2..N hit the
    cache (~10% of input price) instead of re-paying for the body.
    Anthropic's ephemeral cache lasts ~5 min — enough for a per-paper
    extraction batch. The variable per-cell instruction stays uncached.
    """
    try:
        from litellm import completion
    except ImportError:
        sys.exit("litellm not installed. Run: pip install litellm")

    body_trim = body[:LLM_MAX_BODY_CHARS]
    truncated = " [body truncated]" if len(body) > LLM_MAX_BODY_CHARS else ""

    type_line = f"Type: {type_str}" if type_str else "Type: text"
    scale_line = render_scale_for_prompt(scale_dict)
    spec_block = type_line
    if scale_line:
        spec_block += "\n" + scale_line

    system_block = (
        "You are extracting one piece of data from a scientific paper for a "
        "systematic review.\n\n"
        "Output rules:\n"
        "- Return ONLY the extracted value as plain text (no preamble, no JSON, no quotes).\n"
        '- For quantitative fields, quote the verbatim value with units (e.g. "12.4 ± 3.1", "p<0.001").\n'
        "- For ordinal/nominal fields with allowed values, return EXACTLY one of those values.\n"
        '  - If the scale uses codes (e.g. "0 = low, 1 = some concerns, 2 = high"), return the CODE only.\n'
        "- If the paper does NOT report this field, output exactly: not reported\n"
        "- Never invent values; never guess.\n"
        "- Keep the response under 150 characters."
    )
    body_block_text = f"Paper body{truncated}:\n---\n{body_trim}\n---"
    cell_block_text = (
        f"Field name: {column_name}\n"
        f"{spec_block}\n"
        f"Extraction rule: {instruction}\n\n"
        "Return only the extracted value."
    )

    messages = [
        {
            "role": "user",
            "content": [
                {
                    "type": "text",
                    "text": system_block,
                    "cache_control": {"type": "ephemeral"},
                },
                {
                    "type": "text",
                    "text": body_block_text,
                    "cache_control": {"type": "ephemeral"},
                },
                {"type": "text", "text": cell_block_text},
            ],
        }
    ]

    # LLM_MODEL_FAST takes precedence (intended for grunt-work tools);
    # LLM_MODEL is the global override; default is Haiku.
    model = os.getenv("LLM_MODEL_FAST") or os.getenv("LLM_MODEL", DEFAULT_LLM_MODEL)
    try:
        resp = completion(
            model=model,
            messages=messages,
            max_tokens=LLM_MAX_TOKENS,
        )
        value = resp.choices[0].message.content.strip()
        if value.startswith('"') and value.endswith('"'):
            value = value[1:-1]
        time.sleep(LLM_SLEEP_SEC)
        return value
    except Exception as e:
        return f"<llm-error: {type(e).__name__}>"


# ============================================================================
# I/O — XLSX & CSV
# ============================================================================

def detect_format(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".csv":
        return "csv"
    sys.exit(f"Unsupported format: {suffix}. Use .xlsx or .csv.")


def _read_csv_with_encoding_fallback(path):
    """Open a CSV trying common encodings in order.

    Excel on Windows/Mac defaults to cp1252 (with smart quotes, em-dashes,
    accented characters) — opening such a file with strict UTF-8 raises
    UnicodeDecodeError. We try UTF-8 (with BOM tolerance), then cp1252,
    then latin-1 (cannot fail on any byte sequence).
    """
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                reader = csv.DictReader(f)
                headers = list(reader.fieldnames or [])
                data = list(reader)
            if encoding != "utf-8-sig":
                print(f"  · {Path(path).name}: decoded as {encoding} (not UTF-8 — "
                      f"re-save as UTF-8 in Excel for cleaner future reads)",
                      file=sys.stderr)
            return headers, data
        except UnicodeDecodeError:
            continue
    sys.exit(f"Could not decode {path} with any common encoding "
             f"(utf-8 / cp1252 / latin-1). The file may be corrupt or in "
             f"an exotic encoding. Try re-saving as UTF-8 from Excel.")


def read_table(path):
    fmt = detect_format(path)
    if fmt == "xlsx":
        if not HAS_XLSX:
            sys.exit("openpyxl not installed. Run: pip install openpyxl")
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sys.exit(f"Empty workbook: {path}")
        headers = [str(h) if h is not None else "" for h in rows[0]]
        data = []
        for r in rows[1:]:
            data.append({h: ("" if v is None else v) for h, v in zip(headers, r)})
        return headers, data, fmt
    headers, data = _read_csv_with_encoding_fallback(path)
    return headers, data, fmt


def write_table(path, headers, data, fmt):
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data:
            ws.append([row.get(h, "") for h in headers])
        wb.save(path)
    else:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for row in data:
                writer.writerow({h: row.get(h, "") for h in headers})


# ============================================================================
# Instructions row handling
# ============================================================================

def split_spec_rows(rows, slug_col, forced_instructions_row_idx=None):
    """Pop and return the spec rows (INSTRUCTIONS, TYPE, SCALE) if present.

    Returns (spec_dict, remaining_data_rows) where spec_dict has keys
    "instructions", "types", "scales" — each a dict[col → str] (or None
    if the row is absent).

    If `forced_instructions_row_idx` is set (1-indexed counting from the
    first row after headers), force that row to be treated as
    instructions regardless of its slug column marker. Use for 2-row
    templates that don't carry the legacy INSTRUCTIONS marker.
    """
    spec = {"instructions": None, "types": None, "scales": None}
    out = []
    for i, r in enumerate(rows, start=1):
        if forced_instructions_row_idx is not None and i == forced_instructions_row_idx:
            spec["instructions"] = {k: ("" if v is None else str(v)) for k, v in r.items()}
            continue
        slug = str(r.get(slug_col, "")).strip().upper()
        if slug == INSTRUCTIONS_MARKER:
            spec["instructions"] = {k: ("" if v is None else str(v)) for k, v in r.items()}
            continue
        if slug == TYPE_MARKER:
            spec["types"] = {k: ("" if v is None else str(v)) for k, v in r.items()}
            continue
        if slug == SCALE_MARKER:
            spec["scales"] = {k: ("" if v is None else str(v)) for k, v in r.items()}
            continue
        out.append(r)
    return spec, out


# ============================================================================
# Template analysis (--analyze mode)
# ============================================================================

_CATEGORICAL_PIPE_RE = re.compile(r"\s\|\s")
_CATEGORICAL_CODED_RE = re.compile(r"\b\d+\s*=\s*\w")
_CATEGORICAL_COMMA_LIST_RE = re.compile(r"^[A-Za-z][\w\-]*(?:\s*,\s*[A-Za-z][\w\-]*){1,}$")
_UNIT_HINT_RE = re.compile(r"^\(.+\)$")


def classify_instruction(text):
    """Classify a column instruction so the orchestrator can ask the right
    clarifying question if needed.

    Returns a dict with:
      kind            — categorical | nl | empty | type_hint
      inferred_type   — quantitative | ordinal | nominal | text | unknown
      allowed_values  — list[str] if kind == categorical, else None
    """
    s = (text or "").strip()
    if not s:
        return {"kind": "empty", "inferred_type": "unknown", "allowed_values": None}

    if _CATEGORICAL_CODED_RE.search(s):
        pairs = [p.strip() for p in re.split(r"[;,]", s) if "=" in p]
        return {"kind": "categorical", "inferred_type": "ordinal", "allowed_values": pairs}
    if _CATEGORICAL_PIPE_RE.search(s):
        values = [v.strip() for v in s.split("|") if v.strip()]
        return {"kind": "categorical", "inferred_type": "nominal", "allowed_values": values}
    if _CATEGORICAL_COMMA_LIST_RE.match(s):
        values = [v.strip() for v in s.split(",") if v.strip()]
        return {"kind": "categorical", "inferred_type": "nominal", "allowed_values": values}
    if _UNIT_HINT_RE.match(s):
        return {"kind": "type_hint", "inferred_type": "quantitative", "allowed_values": None}

    word_count = len(s.split())
    inferred = "text" if word_count >= 3 else "unknown"
    return {"kind": "nl", "inferred_type": inferred, "allowed_values": None}


def analyze_template_json(template_path, slug_col, forced_instructions_row_idx):
    """Read template, classify each column, return a JSON-serializable analysis."""
    headers, rows, fmt = read_table(template_path)
    if slug_col not in headers:
        return {
            "error": f"Template missing the '{slug_col}' column.",
            "headers": headers,
        }
    spec, data = split_spec_rows(rows, slug_col, forced_instructions_row_idx)
    instructions = spec["instructions"] or {}
    types = spec["types"] or {}
    scales = spec["scales"] or {}

    cols_info = []
    for h in headers:
        if h == slug_col:
            continue
        instr = (instructions.get(h) or "").strip()
        classification = classify_instruction(instr)
        legacy_type = (types.get(h) or "").strip() if types else None
        legacy_scale = (scales.get(h) or "").strip() if scales else None
        cols_info.append({
            "name": h,
            "instruction": instr,
            "kind": classification["kind"],
            "inferred_type": classification["inferred_type"],
            "allowed_values": classification["allowed_values"],
            "legacy_type": legacy_type or None,
            "legacy_scale": legacy_scale or None,
        })

    if types or scales:
        fmt_detected = "legacy-4row"
    elif forced_instructions_row_idx is not None:
        fmt_detected = "forced-2row"
    elif instructions:
        fmt_detected = "instructions-marker-only"
    else:
        fmt_detected = "no-spec"

    return {
        "template_path": str(template_path),
        "format_detected": fmt_detected,
        "n_columns": len(cols_info),
        "n_data_rows": len(data),
        "slug_column": slug_col,
        "columns": cols_info,
    }


# ============================================================================
# --from-source helper
# ============================================================================

def slugs_from_cites(source_slug, sources):
    src = next((s for s in sources if s["slug"] == source_slug), None)
    if not src:
        sys.exit(f"Source not found: {source_slug}")
    cited = {str(d).lower() for d in (src["fm"].get("cites") or [])}
    out = []
    for s in sources:
        d = (s["fm"].get("doi") or "").strip().lower()
        if d and d in cited:
            out.append(s["slug"])
    return sorted(out)


# ============================================================================
# Coded transformation (--coded output mode)
# ============================================================================

_NUMERIC_RE = re.compile(r"-?\d+(?:[\.,]\d+)?")
_SENTINEL_VALUES = {"", "not reported", "instruction_missing", "n/a", "na", "none"}


def code_value(value, column_info):
    """Reduce a verbatim value to the strict form per the column's instruction.

    The "detailed" output keeps the LLM's verbatim answer (with units, quotes,
    qualifiers like "± SD"). The "coded" output strips this to the strict
    publication-ready / analysis-ready form per the column's classification:

      - categorical (no codes)   → canonical allowed-value label (case-fixed)
      - categorical (coded)      → the integer code only ("0", "1", …)
      - quantitative type_hint   → numeric portion only (units stripped)
      - nl / type_hint / unknown → verbatim (no transformation)

    Sentinels ("", "not reported", "INSTRUCTION_MISSING") pass through unchanged.
    """
    if value is None:
        return ""
    v = str(value).strip()
    if not v or v.lower() in _SENTINEL_VALUES:
        return v

    kind = column_info.get("kind", "nl")
    allowed = column_info.get("allowed_values") or []

    if kind == "categorical" and allowed:
        coded_map = {}
        for item in allowed:
            m = re.match(r"\s*(\d+)\s*=\s*(.+)$", str(item))
            if m:
                coded_map[m.group(2).strip().lower()] = m.group(1).strip()

        if coded_map:
            for code in coded_map.values():
                if v == code:
                    return code
            v_lower = v.lower()
            for label, code in coded_map.items():
                if v_lower == label or v_lower.startswith(label):
                    return code
            return ""

        v_lower = v.lower()
        for item in allowed:
            if v_lower == str(item).lower():
                return str(item)
        for item in allowed:
            if str(item).lower() in v_lower:
                return str(item)
        return ""

    if kind == "type_hint":
        m = _NUMERIC_RE.search(v)
        if m:
            return m.group(0).replace(",", ".")
        return ""

    return v


def derive_coded_output_path(detailed_path):
    p = Path(detailed_path)
    stem = p.stem
    if stem.endswith("-filled"):
        stem = stem[: -len("-filled")] + "-coded"
    elif stem.endswith("-detailed"):
        stem = stem[: -len("-detailed")] + "-coded"
    else:
        stem = stem + "-coded"
    return str(p.with_name(stem + p.suffix))


# ============================================================================
# Project folder convention
# ============================================================================

def resolve_project_paths(project_dir):
    """Map a project folder to its canonical artifacts.

    Convention:
      <project>/
      ├── contexte.md            # project scope (narrative)
      ├── instructions.md        # agent-authored extraction spec (per column)
      ├── template.xlsx          # 2-row template (slug + instruction)
      ├── articles/              # sources to extract (links or copies)
      └── output/
          ├── extraction-detailed.xlsx
          └── extraction-coded.xlsx

    All artifacts are optional except `template.{xlsx,csv}`. Missing
    ones are returned as absolute paths so the caller can create them.
    """
    p = Path(project_dir).resolve()
    if not p.is_dir():
        sys.exit(f"--project: {project_dir} is not a directory.")

    template = None
    for cand in (p / "template.xlsx", p / "template.csv"):
        if cand.exists():
            template = cand
            break
    if not template:
        sys.exit(f"--project: no template.xlsx or template.csv found in {p}.")

    output_dir = p / "output"
    return {
        "root": p,
        "context_md": p / "contexte.md",
        "instructions_md": p / "instructions.md",
        "template": template,
        "articles_dir": p / "articles",
        "output_dir": output_dir,
        "detailed_output": output_dir / f"extraction-detailed{template.suffix}",
        "coded_output": output_dir / f"extraction-coded{template.suffix}",
    }



DEFAULT_SR_COLUMNS = [
    "slug", "first_author", "year", "journal", "doi",
    "design", "country",
    "n_total", "n_intervention", "n_control",
    "age_mean", "sex_pct_female", "population", "chronicity", "baseline_fm",
    "intervention", "intervention_subfamily",
    "n_sessions", "session_duration_min",
    "primary_outcome_delta", "p_value", "effect_size", "confidence_interval",
    "adverse_events", "trial_registration",
    "risk_of_bias", "notes",
]

DEFAULT_SR_TYPES = {
    "slug": "text",
    "first_author": "text",
    "year": "quantitative",
    "journal": "text",
    "doi": "text",
    "design": "nominal",
    "country": "text",
    "n_total": "quantitative",
    "n_intervention": "quantitative",
    "n_control": "quantitative",
    "age_mean": "quantitative",
    "sex_pct_female": "quantitative",
    "population": "text",
    "chronicity": "nominal",
    "baseline_fm": "quantitative",
    "intervention": "nominal",
    "intervention_subfamily": "text",
    "n_sessions": "quantitative",
    "session_duration_min": "quantitative",
    "primary_outcome_delta": "quantitative",
    "p_value": "quantitative",
    "effect_size": "quantitative",
    "confidence_interval": "text",
    "adverse_events": "text",
    "trial_registration": "text",
    "risk_of_bias": "ordinal",
    "notes": "text",
}

DEFAULT_SR_SCALES = {
    "year": "(YYYY)",
    "design": "RCT, cohort, cross-sectional, case-control, case-series, non-randomized trial, simulation",
    "n_total": "(integer)",
    "n_intervention": "(integer)",
    "n_control": "(integer)",
    "age_mean": "(years, mean ± SD if reported)",
    "sex_pct_female": "(0-100, percentage)",
    "chronicity": "acute, subacute, chronic",
    "baseline_fm": "(0-66, FM-UE, mean ± SD)",
    "intervention": "BCI, TMS, tDCS, mirror, robot, mental-practice, physio, combined",
    "n_sessions": "(integer)",
    "session_duration_min": "(minutes)",
    "p_value": "(0 to 1, verbatim e.g. 'p<0.01')",
    "risk_of_bias": "0=low, 1=some concerns, 2=high",
}

DEFAULT_SR_INSTRUCTIONS = {
    "slug": "wiki source page basename (e.g. 'cervera-2020')",
    "first_author": "Family name of the first author only",
    "year": "Publication year (4 digits)",
    "journal": "Full journal name",
    "doi": "DOI in canonical form (e.g. 10.1016/j.neubiorev.2012.10.003)",
    "design": "RCT | cohort | cross-sectional | case-control | case-series | non-randomized trial",
    "country": "Country (or countries, semicolon-separated) where the study was conducted",
    "n_total": "Total N analyzed (post-dropout, intention-to-treat preferred)",
    "n_intervention": "N participants in the active/experimental/intervention arm",
    "n_control": "N participants in the control/sham/placebo arm",
    "age_mean": "Mean age of all participants in years (with SD if reported, e.g. '62.4 ± 11.2')",
    "sex_pct_female": "Percentage of female participants (0–100)",
    "population": "Brief description of the population (e.g. 'chronic stroke patients with moderate hemiparesis, FM-UE 25-50')",
    "chronicity": "Time post-stroke or chronicity category (acute / subacute / chronic), with mean if reported",
    "baseline_fm": "Baseline Fugl-Meyer Upper Extremity score, mean ± SD if reported",
    "intervention": "Principal intervention family (BCI, TMS, tDCS, mirror, robot, mental-practice, physio, combined)",
    "intervention_subfamily": "Specific paradigm (e.g. MI-BCI, AO-BCI, hybrid, rTMS-1Hz, rTMS-10Hz, iTBS, cTBS)",
    "n_sessions": "Total number of intervention sessions",
    "session_duration_min": "Duration per session in minutes",
    "primary_outcome_delta": "Mean change in the PRIMARY outcome from baseline to end-of-treatment, in the intervention arm, with units (verbatim)",
    "p_value": "p-value of the primary between-group comparison (verbatim, e.g. 'p<0.01' or 'p=0.034')",
    "effect_size": "Effect size of primary outcome (Cohen's d, Hedges' g, η², or mean difference with units)",
    "confidence_interval": "95% CI of the primary effect (e.g. '2.1 to 5.4' or '0.32 to 1.07')",
    "adverse_events": "Number of adverse events reported (intervention arm, then control arm), or 'none reported' / 'not reported'",
    "trial_registration": "Trial registration ID (NCT, ISRCTN, EudraCT, CTRI, ChiCTR), or empty if not registered",
    "risk_of_bias": "Cochrane RoB 2 overall judgment (low / some concerns / high) with the main domain of concern (e.g. 'some concerns – blinding of outcome assessor')",
    "notes": "Free-text notes — manual fill",
}


# ============================================================================
# Main
# ============================================================================

def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("template", nargs="?",
                    help="Excel (.xlsx) or CSV template to populate.")
    ap.add_argument("--output", "-o",
                    help="Write to this path. Default: <template-stem>-filled<ext> "
                         "next to the template (the template stays as a reusable spec). "
                         "Use --in-place to overwrite the template instead.")
    ap.add_argument("--in-place", action="store_true",
                    help="Overwrite the template in place (legacy behavior). "
                         "Default: write to a sibling <stem>-filled file.")
    ap.add_argument("--from-source",
                    help="Pre-fill a NEW template from a source's cites: "
                         "(use with --output).")
    ap.add_argument("--columns",
                    help="Comma-separated column names for --from-source "
                         "(default: a sensible SR set).")
    ap.add_argument("--no-spec", action="store_true",
                    help="With --from-source: skip the INSTRUCTIONS/TYPE/SCALE rows.")
    ap.add_argument("--slug-column", default="slug",
                    help="Name of the column listing source slugs (default: 'slug').")
    ap.add_argument("--llm", action="store_true",
                    help="Fall back to LLM (litellm) for cells still empty "
                         "after frontmatter + regex, using the INSTRUCTIONS "
                         "row to drive extraction.")
    ap.add_argument("--analyze", action="store_true",
                    help="Read the template, classify each column's instruction, "
                         "output a JSON analysis on stdout. Does not extract. "
                         "Use this from the orchestrator agent to drive the "
                         "comprehension-debrief gate.")
    ap.add_argument("--instructions-row", type=int, default=None,
                    help="Treat row N (1-indexed, counting from the first row "
                         "after the headers) as the instructions row. Use for "
                         "2-row templates that don't carry the legacy "
                         "INSTRUCTIONS marker in the slug column. Default: "
                         "auto-detect via the legacy markers.")
    ap.add_argument("--coded", action="store_true",
                    help="In addition to the detailed output, write a 'coded' "
                         "sibling file with strict per-instruction format "
                         "(canonical category labels, codes only for ordinal-"
                         "coded scales, units stripped for quantitative). "
                         "Use this for the publication-ready / R-ready table.")
    ap.add_argument("--project",
                    help="Operate on a literature-review project folder. "
                         "Expects template.xlsx (or .csv) at root; writes "
                         "outputs to <project>/output/extraction-{detailed,coded}.<ext>. "
                         "Implies --coded. Optional contexte.md / instructions.md "
                         "at the project root are read/written by the slash command "
                         "(/wiki-extract-table), not by this script.")
    args = ap.parse_args()

    project = None
    if args.project:
        project = resolve_project_paths(args.project)
        if args.template:
            sys.exit("--project and a positional template are mutually exclusive.")
        args.template = str(project["template"])
        if not args.output:
            args.output = str(project["detailed_output"])
        args.coded = True
        project["output_dir"].mkdir(parents=True, exist_ok=True)

    # ---------- Analyze mode (no wiki access needed) ----------
    if args.analyze:
        if not args.template:
            sys.exit("--analyze requires a template path.")
        result = analyze_template_json(args.template, args.slug_column, args.instructions_row)
        print(json.dumps(result, indent=2, ensure_ascii=False))
        return

    sources = load_sources()
    if not sources:
        sys.exit("No sources in wiki/sources/")
    sources_index = {s["slug"]: s for s in sources}

    # ---------- Pre-fill mode ----------
    if args.from_source:
        if not args.output:
            sys.exit("--from-source requires --output")
        slugs = slugs_from_cites(args.from_source, sources)
        cols = (args.columns.split(",") if args.columns else DEFAULT_SR_COLUMNS)
        cols = [c.strip() for c in cols]
        if args.slug_column not in cols:
            cols = [args.slug_column] + cols

        rows = []
        if not args.no_spec:
            instr_row = {c: DEFAULT_SR_INSTRUCTIONS.get(c, "<add your extraction rule here>") for c in cols}
            instr_row[args.slug_column] = INSTRUCTIONS_MARKER
            type_row = {c: DEFAULT_SR_TYPES.get(c, "text") for c in cols}
            type_row[args.slug_column] = TYPE_MARKER
            scale_row = {c: DEFAULT_SR_SCALES.get(c, "") for c in cols}
            scale_row[args.slug_column] = SCALE_MARKER
            rows.extend([instr_row, type_row, scale_row])
        for slug in slugs:
            row = {c: "" for c in cols}
            row[args.slug_column] = slug
            rows.append(row)

        fmt = detect_format(args.output)
        write_table(args.output, cols, rows, fmt)
        suffix = " (with INSTRUCTIONS / TYPE / SCALE rows)" if not args.no_spec else ""
        print(f"  ✓ {args.output} pre-filled with {len(slugs)} slugs from "
              f"{args.from_source}'s cites{suffix}")
        return

    # ---------- Fill mode ----------
    if not args.template:
        sys.exit("Provide a template file (or use --from-source).")
    headers, rows, fmt = read_table(args.template)

    if args.slug_column not in headers:
        sys.exit(f"Template must contain a '{args.slug_column}' column. "
                 f"Got: {headers}")

    spec, data = split_spec_rows(rows, args.slug_column, args.instructions_row)
    instructions = spec["instructions"]
    types = spec["types"] or {}
    scales = spec["scales"] or {}

    parsed_scales = {col: parse_scale(scales.get(col, "")) for col in headers}

    detected = []
    if instructions:
        detected.append(f"INSTRUCTIONS ({sum(1 for v in instructions.values() if v.strip())})")
    if spec["types"]:
        detected.append(f"TYPE ({sum(1 for v in types.values() if v.strip())})")
    if spec["scales"]:
        detected.append(f"SCALE ({sum(1 for v in scales.values() if v.strip())})")
    if detected:
        print(f"  · spec rows detected: {', '.join(detected)}", file=sys.stderr)
    if args.llm and not instructions:
        print("  ! --llm requested but no INSTRUCTIONS row found. "
              "Add a row with slug=INSTRUCTIONS to drive LLM extraction.",
              file=sys.stderr)

    llm_cache = load_llm_cache() if args.llm else {}

    payload_cols = [h for h in headers if h != args.slug_column]
    method_count = {"frontmatter": 0, "regex": 0, "llm": 0, "manual": 0, "empty": 0, "invalid": 0}
    row_status = {"complete": 0, "partial": 0, "empty": 0, "not_found": 0}

    for row in data:
        slug = (str(row.get(args.slug_column) or "")).strip()
        if not slug:
            continue
        src = sources_index.get(slug)
        if not src:
            row_status["not_found"] += 1
            continue

        for col in payload_cols:
            existing = str(row.get(col) or "").strip()
            if existing:
                method_count["manual"] += 1
                continue

            col_norm = normalize_col(col)
            col_type = (types.get(col) or "").strip().lower() if types else ""
            col_scale = parsed_scales.get(col)

            # 1. Frontmatter
            v = apply_fm_map(col_norm, src)
            if v not in (None, ""):
                ok, normalized, warn = validate_value(str(v), col_type, col_scale)
                row[col] = normalized
                if ok:
                    method_count["frontmatter"] += 1
                else:
                    method_count["invalid"] += 1
                    print(f"  ! [{slug}/{col}] frontmatter value didn't validate: {warn}",
                          file=sys.stderr)
                continue

            # 2. Body regex
            v = extract_from_body(col_norm, src["body"])
            if v not in (None, ""):
                ok, normalized, warn = validate_value(str(v), col_type, col_scale)
                row[col] = normalized
                if ok:
                    method_count["regex"] += 1
                else:
                    method_count["invalid"] += 1
                    print(f"  ! [{slug}/{col}] regex match didn't validate: {warn}",
                          file=sys.stderr)
                continue

            # 3. LLM (if instruction provided)
            if args.llm and instructions:
                instr = (instructions.get(col) or "").strip()
                if instr and instr != "<add your extraction rule here>":
                    cache_key = f"{slug}::{col_norm}::{hash((instr, col_type, str(col_scale)))}"
                    if cache_key in llm_cache:
                        value = llm_cache[cache_key]
                    else:
                        # Section windowing: send only the relevant IMRAD
                        # section (Methods / Results / Discussion / Intro)
                        # instead of the full body. Cuts ~50% of input
                        # tokens for clinical fields.
                        windowed_body, section_used = window_body(src["body"], col_norm)
                        value = llm_extract(col, instr, windowed_body, col_type, col_scale)
                        llm_cache[cache_key] = value
                    ok, normalized, warn = validate_value(value, col_type, col_scale)
                    row[col] = normalized
                    if not ok:
                        method_count["invalid"] += 1
                        print(f"  ! [{slug}/{col}] LLM output didn't validate: {warn} (raw: {value!r})",
                              file=sys.stderr)
                    elif normalized.lower() == "not reported":
                        method_count["empty"] += 1
                    else:
                        method_count["llm"] += 1
                    continue

            method_count["empty"] += 1

        # Per-row status
        n_filled = sum(1 for c in payload_cols if str(row.get(c) or "").strip())
        if n_filled == len(payload_cols):
            row_status["complete"] += 1
        elif n_filled > 0:
            row_status["partial"] += 1
        else:
            row_status["empty"] += 1

    if args.llm:
        save_llm_cache(llm_cache)

    if args.output:
        out = args.output
    elif args.in_place:
        out = args.template
    else:
        tpl = Path(args.template)
        out = str(tpl.with_name(tpl.stem + "-filled" + tpl.suffix))

    # Re-include the spec rows at the top of the output (preserve user's edits)
    final_rows = []
    if instructions:
        final_rows.append(instructions)
    if spec["types"]:
        final_rows.append(spec["types"])
    if spec["scales"]:
        final_rows.append(spec["scales"])
    final_rows.extend(data)
    write_table(out, headers, final_rows, fmt)

    if out == args.template:
        print(f"  ✓ {out} updated (in-place).")
    else:
        print(f"  ✓ {out} written  (template preserved at {args.template}).")

    # Coded sibling output — strict per-instruction format
    if args.coded:
        col_classifications = {}
        for h in headers:
            if h == args.slug_column:
                continue
            instr = (instructions.get(h) or "").strip() if instructions else ""
            col_classifications[h] = {
                "instruction": instr,
                **classify_instruction(instr),
            }

        coded_data = []
        for row in data:
            coded_row = {}
            for h in headers:
                if h == args.slug_column:
                    coded_row[h] = row.get(h, "")
                else:
                    coded_row[h] = code_value(row.get(h, ""), col_classifications[h])
            coded_data.append(coded_row)

        if args.project:
            coded_out = str(project["coded_output"])
        else:
            coded_out = derive_coded_output_path(out)

        coded_final = list(final_rows[: -len(data)]) + coded_data
        write_table(coded_out, headers, coded_final, fmt)
        print(f"  ✓ {coded_out} written  (coded — strict per-instruction format).")
    print(f"\nPer-row status:")
    for k, v in row_status.items():
        print(f"  {k}: {v}")
    print(f"\nPer-cell method:")
    for k, v in method_count.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
